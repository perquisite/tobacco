# -*- coding: utf-8 -*-
import shutil
import sys
import os

import openpyxl
import win32com
from openpyxl.styles import Alignment
from win32com.client import Dispatch

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)

from yancaoRegularDemo.Resource.Multi_Table.MultiTableProcessor import MultiTableProcessor
from yancaoRegularDemo.Resource.Multi_Table.MultiTableProcessor import wait_time
from yancaoRegularDemo.Resource.tangyuhao import Precessor1
from yancaoRegularDemo.Resource.tools.get_pictures import get_pictures_single
import re
import time
import xlwt
import pythoncom

def anjuan_api(input_path, export_dir):
    try:
        pythoncom.CoInitialize()
        word = Dispatch('Word.Application')
        pythoncom.CoInitialize()
        word.Documents.Close()
    except:
        print("关闭word进程")
    # 这里先把文件夹里的doc文件替换为docx
    # 注意：若文件夹里包含相同名字但后缀分别为doc和docx的两个文件，则docx的文件会被doc的文件转化后的docx所覆盖！
    try:
        replace_doc_to_docx(input_path)
    except Exception as e:
        print("replace_doc_to_docx函数出错：" + str(e.args))
    input_path = input_path.replace("\\", "/")
    export_dir = export_dir.replace("\\", "/")
    input_dir_dictionary = {}
    # 果如有picture文件夹，删掉
    if os.path.exists(os.path.join(input_path, "picture")):
        shutil.rmtree(os.path.join(input_path, "picture"))
    # 先看是多文件夹，还是直接存放的一套文书docx
    multi = True
    # 判断多文件夹还是只含一套的docx
    for root, dirs, files in os.walk(input_path):
        if len(dirs) >= 1 and 'data' not in dirs:  # 含多文件夹
            input_dir_dictionary = get_input_dir_dictionary(input_path)
            # print(input_dir_dictionary)
            # 提取文档中的图片
            get_pictures(root, dirs)
            break
        else:
            multi = False
            break
    if not multi:
        return one_Processor(root, files, export_dir)
    else:
        return multi_Processor(input_dir_dictionary, export_dir)


def doc_to_docx(file):
    if file[-3:] == "doc":
        word = win32com.client.Dispatch("Word.Application")
        # print(file)
        try:
            doc = word.Documents.Open(file)
            doc.SaveAs(file + "x", 12)
            doc.Close()
        except Exception as e:
            print("文件" + file + "在转为docx时出现错误——" + str(e.args))
        file_path = file + "x"
        # print(file_path)
        return file_path
    else:
        file_path = file
        return False


def replace_doc_to_docx(input_path):
    file_name_list = os.listdir(input_path)
    for file in file_name_list:
        if file[-4:] == ".doc":
            # print(input_path + "\\" + file)
            if doc_to_docx(input_path + "\\" + file):
                os.remove(input_path + "\\" + file)


def one_Processor(root, files, export_dir):
    get_pictures_single(root)  # 提取文档内的图片
    # print(files)
    result_dic = {}
    for f in files:
        info_list = []
        file_path = root + "/" + f
        # print(file_path)
        time.sleep(wait_time)
        try:
            temp = Precessor1(file_path, export_dir, True).action()
        except Exception as e:
            print(f + ' 未成功处理，已跳过。遇到错误：' + str(e.args))
            continue
        info_list += temp
        contract_check_result = ["", []]
        contract_check_result[0] = f
        contract_check_result[1] = temp
        result_dic[f] = info_list
    # 先拷贝一份模板到输出文件夹
    excel_path = export_dir + '\\' + '案卷审查报告.xlsx'
    # print(excel_path)
    shutil.copyfile(rootPath+'\\'+'yancaoRegularDemo'+'\\'+'案卷报告模板.xlsx', excel_path)
    write_excel_single(result_dic, root, excel_path)
    return result_dic


def multi_Processor(input_dir_dictionary, export_dir):
    m = MultiTableProcessor(input_dir_dictionary, 20, export_dir, True, None, None)
    contract_check_result = m.check()
    all_info = m.get_all_info()
    write_excel_multi(contract_check_result, export_dir, all_info)
    return m.api_info


def write_excel_multi(contract_check_result, export_dir, all_info):
    style_common = xlwt.XFStyle()
    # 判断输出文件是否存在，若不存在则创建文件文件
    flag = 0
    #time0 = time.strftime("%Y-%m-%d_%H_%M_%S", time.localtime())
    file_p = os.path.join(export_dir + '/' + "案卷审查结果表格_" + ".xls")
    if flag == 0:
        file_all = []
        for root, dirs, files in os.walk(export_dir):
            for f in files:
                file_all.append(os.path.join(export_dir + '/', f))
        if file_p not in file_all:
            doc0 = xlwt.Workbook(encoding='utf-8')
            worksheet = doc0.add_sheet('My Worksheet')
            doc0.save(file_p)
        else:
            print("已存在结果文件")
        flag = 1
    # 写入表格
    # if have_dir:
    doc1 = xlwt.Workbook(file_p)
    # get_all_inf = Multi.get_all_info()
    worksheet_all = doc1.add_sheet("总览")
    first_col = worksheet_all.col(0)  # xlwt中是行和列都是从0开始计算的
    sec_col = worksheet_all.col(1)
    third_col = worksheet_all.col(2)
    forth_col = worksheet_all.col(3)

    first_col.width = 256 * 40
    sec_col.width = 256 * 20
    third_col.width = 256 * 20
    forth_col.width = 256 * 40

    line = 0

    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.bold = True
    font.height = 18 * 20
    style.font = font

    worksheet_all.write(0, 0, "案卷名称", style)
    worksheet_all.write(0, 1, "实际审查项数量", style)
    worksheet_all.write(0, 2, "错误项数量", style)
    worksheet_all.write(0, 3, "需要人工审查的数量", style)
    line = line + 1
    for item in all_info:
        worksheet_all.write(line, 0, item[0], style_common)
        worksheet_all.write(line, 1, item[1], style_common)
        worksheet_all.write(line, 2, item[2], style_common)
        worksheet_all.write(line, 3, item[3], style_common)
        line = line + 1

    dir_name0 = None
    for item in contract_check_result:
        if isinstance(item, str):
            if "self." not in item:
                pass
            else:
                contract_check_result.remove(item)
                continue
        elif isinstance(item, list):
            item = [i for i in item if i[:5] != 'self.']
        if "-----------------------------------\n正在审查文件：" in item:
            pattern = r"-----------------------------------\n正在审查文件：(.*)"
            t = re.findall(pattern, item)[0]  # t代表文件名

            end_pos = t.rfind('/') - 1
            dir_name = t[t.rfind('/', 1, end_pos) + 1:t.rfind('/', 1)]  # dir_name代表文件所在的文件夹
            if dir_name0 == None:
                line = 0
                dir_name0 = dir_name  # 储存上一个文件夹
                worksheet = doc1.add_sheet(dir_name)
                first_col = worksheet.col(0)  # xlwt中是行和列都是从0开始计算的
                sec_col = worksheet.col(1)

                first_col.width = 256 * 70
                sec_col.width = 256 * 100

                style = xlwt.XFStyle()
                font = xlwt.Font()
                font.bold = True
                font.height = 18 * 20
                style.font = font

                worksheet.write(0, 0, '文书路径', style)
                worksheet.write(0, 1, '错误内容', style)
                line = line + 1

            if dir_name != dir_name0:  # 换sheet
                worksheet = doc1.add_sheet(dir_name)
                first_col = worksheet.col(0)  # xlwt中是行和列都是从0开始计算的
                sec_col = worksheet.col(1)

                first_col.width = 256 * 70
                sec_col.width = 256 * 100

                dir_name0 = dir_name
                line = 0

                style = xlwt.XFStyle()
                font = xlwt.Font()
                font.bold = True
                font.height = 18 * 20
                style.font = font

                worksheet.write(0, 0, '文书路径', style)
                worksheet.write(0, 1, '错误内容', style)
                line = line + 1

        elif "正在审查同案由案件比较信息：" in item:
            line = 0
            worksheet_compare = doc1.add_sheet("比较结果")
            first_col = worksheet_compare.col(0)  # xlwt中是行和列都是从0开始计算的
            sec_col = worksheet_compare.col(1)

            first_col.width = 256 * 70
            sec_col.width = 256 * 200

            style = xlwt.XFStyle()
            font = xlwt.Font()
            font.bold = True
            font.height = 18 * 20
            style.font = font

            worksheet_compare.write(0, 0, "比较项名称", style)
            worksheet_compare.write(0, 1, "比较结果", style)
            line = line + 1


        elif "案件处理审批表_.docx\n" in item:
            worksheet_compare.write(line, 0, item, style_common)

        elif "同案由案件" in item:
            worksheet_compare.write(line, 1, item, style_common)
            line = line + 1

        else:
            for i in item:
                if "×" in i:
                    i = i[2:]
                worksheet.write(line, 0, t, style_common)
                worksheet.write(line, 1, i, style_common)
                line = line + 1

    doc1.save(file_p)


"""
def write_excel_single(filename, contract_check_result, export_dir):
    style_common = xlwt.XFStyle()
    # time0 = time.strftime("%Y-%m-%d_%H_%M_%S", time.localtime())
    # file_p = os.path.join(export_dir + '/' + "案卷审查结果表格_" + time0 + ".xls")
    fn = filename.strip('.docx').strip('.doc')
    file_p = os.path.join(export_dir + '/' + fn + "案卷审查结果表格_" + ".xls")
    doc1 = xlwt.Workbook(file_p)
    for item in contract_check_result:
        if isinstance(item, str):
            if "self." not in item:
                pass
            else:
                contract_check_result.remove(item)
        elif isinstance(item, list):
            item = [i for i in item if i[:5] != 'self.']
        if ".docx" in item:
            line = 0
            worksheet = doc1.add_sheet(
                item[item.rfind('/', 1) + 1:item.rfind('.', 1)])
            first_col = worksheet.col(0)  # xlwt中是行和列都是从0开始计算的
            sec_col = worksheet.col(1)

            first_col.width = 256 * 70
            sec_col.width = 256 * 100
        else:
            for i in item:
                worksheet.write(line, 0, i, style_common)
                line = line + 1
    doc1.save(file_p)
"""


def get_pictures(root, dirs):
    for d in dirs:
        path = root + "/" + d
        get_pictures_single(path)


def get_input_dir_dictionary(input_path):
    file_paths = []
    input_dir_dictionary = {}
    for root, dirs, files in os.walk(input_path):
        for d in dirs:
            p = input_path + "/" + d
            for root, dirs, files in os.walk(p):
                for f in files:
                    file_path = p + "/" + f
                    file_paths.append(file_path)
            input_dir_dictionary[p] = file_paths
            file_paths = []
    return input_dir_dictionary


def write_excel_single(result_dic, root_path, excel_path):
    # 先拷贝一份模板sheet
    # i = root_path.rfind('/') + 1
    # anjuan_name = root_path[i:]
    # 要过滤掉anjuan_name中的'*', ':', '/', '\\', '?', '[', ']'
    # anjuan_name = anjuan_name.replace("*", "").replace(":", "").replace("/", "").replace("\\", "").replace("?", "").replace("[", "").replace("]", "")
    anjuan_name = '案卷审查结果'
    wb = openpyxl.load_workbook(excel_path)
    sheet1 = wb['Sheet1']
    new_sheet = wb.copy_worksheet(sheet1)
    new_sheet.title = anjuan_name
    wb.remove(sheet1)
    wb.save(excel_path)
    # 写题头
    sheet = wb[anjuan_name]
    sheet.cell(row=3, column=2, value=anjuan_name)
    time0 = time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime())
    sheet.cell(row=3, column=5, value=time0)
    # 写入正式内容
    start_row = 5
    no = 1
    for item in result_dic:
        # 先写文书名称
        align = Alignment(horizontal='center', vertical='center')
        wenshu_name = str(no) + '、' + item.strip('.docx').strip('doc')
        sheet.cell(row=start_row, column=2, value=wenshu_name).alignment = align
        # 再写审查结果
        start_row += 1
        for record in result_dic[item]:
            if 'self.' in record or '：正确。' in record:
                continue
            if '不存在表\n-----------------------------------' in record:
                sheet.cell(row=start_row, column=2, value='【该文书未被审查，原因可能如下：1.文书不存在 2.文书文件名不规范，无法识别 3.文书仅被用于对比，无需审查（如询问（调查）通知书）】')
                continue
            lst = split_info(record)
            sheet.cell(row=start_row, column=1, value=lst[0]).alignment = align
            sheet.cell(row=start_row, column=2, value=lst[1])
            start_row += 1
        start_row += 2
        no += 1

    wb.save(excel_path)
    wb.close()


def split_info(info):
    split_loc = info.find('：')
    info = [info[:split_loc+1], info[split_loc+1:]]
    return info


# 命令：python Anjuan_api.py input_path export_path
# 一定要绝对路径，且文件名不能有空格
if __name__ == '__main__':
    args = sys.argv
    input_path = args[1]
    export_dir = args[-1]
    print("输入路径：{}，输出路径：{}，".format(input_path, export_dir))
    # input_path = "C:\\Users\\Xie\\Desktop\\in"
    # export_dir = "C:\\Users\\Xie\\Desktop\\out"
    res = anjuan_api(input_path, export_dir)
    print(res)

